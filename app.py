import streamlit as st
import pandas as pd
import requests
import io
import dataframe_image as dfi
import base64
from datetime import datetime, timedelta
import pytz 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
import os

from PIL import Image, ImageDraw, ImageFont

st.set_page_config(page_title="Gestão Mimo", layout="centered")
# ==========================================
# CONFIGURAÇÕES
# ==========================================
URL_PLANILHA = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSH9lJhzNgDz3x05wnE3lc24YKiUQcn_WTNgxEpsSO2jA36rAwSDfLZUkm1SgE_uoKBXvgx1_8sDTXZ/pub?output=xlsx"

COLUNA_FILTRO_HORA = 'INI' 
COL_PERIODO = 'ENT'           
COL_HORA = 'INI'              
COL_LINHA = 'LINHA'           
COL_EMPRESA = 'CLIENTE'       
COL_PREFIXO = 'FROTA FINAL' 
COL_MOTORISTA = 'MOTORISTA'   

MAPA_LOGOS = {
    "MELI": "logo_meli.png", "MERCADO LIVRE": "logo_meli.png", 
    "AMAZON": "logo_amazon.png", "ADORO": "logo_adoro.png", "AAM": "logo_aam.png"
}

MAPA_GRUPOS = {
    "MELI": "120363000000000000@g.us", "AMAZON": "120363000000000001@g.us", 
    "ADORO": "5511917623237", "AAM": "5511934773679"
}

URL_EVOLUTION = "https://mimo-evolution-api.3sbqz4.easypanel.host/message/sendMedia/teste"
CHAVE_API_EVOLUTION = "429683C4C977415CAAFCCE10F7D57E11"

# ==========================================
# FUNÇÕES DE IMAGEM E ENVIO
# ==========================================
def embutir_logos_na_imagem(img_path, cliente_nome):
    try:
        tabela_img = Image.open(img_path)
        largura_tabela, altura_tabela = tabela_img.size
        altura_cabecalho = 160
        nova_largura = max(largura_tabela, 800) 
        nova_img = Image.new('RGB', (nova_largura, altura_tabela + altura_cabecalho), 'white')
        
        x_tabela = (nova_largura - largura_tabela) // 2
        nova_img.paste(tabela_img, (x_tabela, altura_cabecalho))
        
        draw = ImageDraw.Draw(nova_img)
        texto_titulo = f"PROGRAMAÇÃO - {cliente_nome}"
        try:
            font = ImageFont.truetype("DejaVuSans-Bold.ttf", 32)
        except:
            font = ImageFont.load_default()

        w_texto = draw.textlength(texto_titulo, font=font)
        x_texto = (nova_largura - w_texto) // 2
        draw.text((x_texto, 100), texto_titulo, fill=(255, 0, 0), font=font)
        
        try:
            mimo = Image.open('logo_mimo.png')
            mimo.thumbnail((200, 80))
            nova_img.paste(mimo, (20, 20), mimo if mimo.mode == 'RGBA' else None)
        except: pass
        
        try:
            for chave, arquivo in MAPA_LOGOS.items():
                if chave in cliente_nome:
                    cliente_logo = Image.open(arquivo)
                    cliente_logo.thumbnail((160, 80))
                    nova_img.paste(cliente_logo, (nova_largura - 180, 20), cliente_logo if cliente_logo.mode == 'RGBA' else None)
                    break
        except: pass
        nova_img.save(img_path)
    except Exception as e: print(f"Erro ao montar imagem: {e}")

def enviar_evolution(imagem_path, nome_empresa, data_str, contexto):
    id_grupo = next((v for k, v in MAPA_GRUPOS.items() if k in nome_empresa), None)
    if not id_grupo: return f"⚠️ Destino não configurado: {nome_empresa}"
    if "@c.us" in id_grupo: id_grupo = id_grupo.replace("@c.us", "")

    msg = f"🚌 *Programação de Escala*\n🏢 *Cliente:* {nome_empresa}\n📅 *Data:* {data_str}\n⏱️ *Janela:* {contexto}"
    headers = {"Content-Type": "application/json", "apikey": CHAVE_API_EVOLUTION}

    try:
        with open(imagem_path, 'rb') as f:
            base64_data = base64.b64encode(f.read()).decode('ascii')
        
        payload = {"number": id_grupo, "mediatype": "image", "media": base64_data, "caption": msg}
        resp = requests.post(URL_EVOLUTION, headers=headers, json=payload)
        if resp.status_code in [200, 201]: return "✅ Escala enviada!"
        else: return f"❌ Erro Evolution: {resp.text}"
    except Exception as e: return f"❌ Falha de conexão: {e}"

def gerar_planilha_formatada(df, cliente_id):
    wb = Workbook(); ws = wb.active
    fill_vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True)
    ws.merge_cells('A12:F12')
    ws['A12'] = f"PROGRAMAÇÃO - {cliente_id}"
    ws.append([]); ws.append(["Periodo", "Horas", "Linha", "Empresa", "Prefixo", "Motorista"])
    for col in range(1, 7):
        c = ws.cell(row=14, column=col); c.fill = fill_vermelho; c.font = fonte_branca
    for _, row in df.iterrows():
        ws.append([row.get(COL_PERIODO,''), row.get(COL_HORA,''), row.get(COL_LINHA,''), row.get(COL_EMPRESA,''), row.get(COL_PREFIXO,''), row.get(COL_MOTORISTA,'')])
    ws.column_dimensions['C'].width = 45; ws.column_dimensions['F'].width = 25
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ==========================================
# GATILHO INVISÍVEL DO N8N (A MÁGICA SEGURA)
# ==========================================
query_params = st.query_params

if "api_n8n" in query_params:
    st.title("⚙️ Robô Automático Executando...")
    cliente_alvo = query_params.get("cliente", "").upper()
    horario_alvo = query_params.get("horario", "")
    
    if cliente_alvo and horario_alvo:
        try:
            fuso = pytz.timezone('America/Sao_Paulo')
            agora = datetime.now(fuso).replace(tzinfo=None)
            
            # 1. Matemática da Janela de 2 Horas
            hora_obj = datetime.strptime(horario_alvo, '%H:%M').time()
            inicio_filtro = agora.replace(hour=hora_obj.hour, minute=hora_obj.minute, second=0, microsecond=0)
            fim_filtro = inicio_filtro + timedelta(hours=2)
            
            # 2. Baixa e lê o Google Sheets do dia
            r = requests.get(URL_PLANILHA)
            xls = pd.ExcelFile(r.content)
            nome_aba = agora.strftime("%d%m%Y")
            
            if nome_aba in [a.strip() for a in xls.sheet_names]:
                df_bruto = pd.read_excel(xls, sheet_name=nome_aba, header=None)
                linha_cab = next((i for i, r in df_bruto.iterrows() if any(str(v).strip().upper() == COLUNA_FILTRO_HORA for v in r.values)), None)
                df = df_bruto.iloc[linha_cab + 1:].reset_index(drop=True)
                df.columns = [str(c).strip().upper() for c in df_bruto.iloc[linha_cab]]
                df = df.dropna(subset=[COLUNA_FILTRO_HORA])
                
                def converter_tempo(v):
                    if pd.isna(v): return pd.NaT
                    try:
                        dt = v if hasattr(v, 'hour') else pd.to_datetime(str(v).replace('h', ':').strip())
                        return agora.replace(hour=dt.hour, minute=dt.minute, second=0, microsecond=0)
                    except: return pd.NaT

                df['AUX_TIME'] = df[COLUNA_FILTRO_HORA].apply(converter_tempo)
                
                # 3. Filtra pelo Cliente e pela Janela exata
                df_filtrado = df[
                    (df[COL_EMPRESA].str.contains(cliente_alvo, na=False)) & 
                    (df['AUX_TIME'] >= inicio_filtro) & 
                    (df['AUX_TIME'] <= fim_filtro)
                ].copy()
                
                if not df_filtrado.empty:
                    cols_print = [c for c in [COL_PERIODO, COL_HORA, COL_LINHA, COL_EMPRESA, COL_PREFIXO, COL_MOTORISTA] if c in df_filtrado.columns]
                    style = df_filtrado[cols_print].style.set_properties(**{'background-color': 'white', 'color': 'black', 'border': '1px solid black'})\
                        .set_table_styles([{'selector': 'th', 'props': [('background-color', '#FF0000'), ('color', 'white')]}])
                    
                    img_path = f"auto_{cliente_alvo}.png"
                    dfi.export(style, img_path, table_conversion="matplotlib", max_rows=-1)
                    embutir_logos_na_imagem(img_path, cliente_alvo)
                    
                    msg_janela = f"De {horario_alvo} às {fim_filtro.strftime('%H:%M')}"
                    res = enviar_evolution(img_path, cliente_alvo, agora.strftime('%d/%m/%Y'), msg_janela)
                    st.success(res)
                else:
                    st.warning("Sem viagens para essa janela.")
            else:
                st.error("Aba não encontrada.")
        except Exception as e:
            st.error(f"Erro Crítico: {e}")
    st.stop() # Finaliza aqui. O robô invisível não mostra botões!

# ==========================================
# INTERFACE PRINCIPAL (VISUAL HUMANO)
# ==========================================
st.title("Gerador de Escalas por Cliente 🚌⏳")

if 'clientes_processados' not in st.session_state:
    st.session_state.clientes_processados = {}

if st.button("1. Analisar Planilha e Gerar Prévias (Próximas 3h)", type="primary"):
    with st.spinner("Analisando planilha..."):
        try:
            fuso = pytz.timezone('America/Sao_Paulo')
            agora = datetime.now(fuso).replace(tzinfo=None)
            inicio_filtro = agora - timedelta(minutes=20)
            fim_filtro = agora + timedelta(hours=3)
            
            r = requests.get(URL_PLANILHA)
            xls = pd.ExcelFile(r.content)
            nome_aba = agora.strftime("%d%m%Y")

            df_bruto = pd.read_excel(xls, sheet_name=nome_aba, header=None)
            linha_cab = next((i for i, r in df_bruto.iterrows() if any(str(v).strip().upper() == COLUNA_FILTRO_HORA for v in r.values)), None)
            df = df_bruto.iloc[linha_cab + 1:].reset_index(drop=True)
            df.columns = [str(c).strip().upper() for c in df_bruto.iloc[linha_cab]]
            df = df.dropna(subset=[COLUNA_FILTRO_HORA]) 

            def converter_tempo(v):
                if pd.isna(v): return pd.NaT
                try:
                    dt = v if hasattr(v, 'hour') else pd.to_datetime(str(v).replace('h', ':').strip())
                    return agora.replace(hour=dt.hour, minute=dt.minute, second=0, microsecond=0)
                except: return pd.NaT

            df['AUX_TIME'] = df[COLUNA_FILTRO_HORA].apply(converter_tempo)
            df_filtrado = df[(df['AUX_TIME'] >= inicio_filtro) & (df['AUX_TIME'] <= fim_filtro)].copy()

            clientes_dict = {}
            for cliente, group_df in df_filtrado.groupby(COL_EMPRESA):
                cliente_nome = str(cliente).strip().upper()
                nome_seguro = cliente_nome.replace("/", "_").replace(":", "_")
                img_path = f"escala_{nome_seguro}.png"
                
                cols_print = [c for c in [COL_PERIODO, COL_HORA, COL_LINHA, COL_EMPRESA, COL_PREFIXO, COL_MOTORISTA] if c in group_df.columns]
                style = group_df[cols_print].style.set_properties(**{'background-color': 'white', 'color': 'black', 'border': '1px solid black'})\
                    .set_table_styles([{'selector': 'th', 'props': [('background-color', '#FF0000'), ('color', 'white')]}])
                
                dfi.export(style, img_path, table_conversion="matplotlib", max_rows=-1)
                embutir_logos_na_imagem(img_path, cliente_nome)
                
                clientes_dict[cliente_nome] = {
                    "img": img_path, 
                    "excel": gerar_planilha_formatada(group_df, cliente_nome), 
                    "data_str": agora.strftime('%d/%m/%Y')
                }
                
            st.session_state.clientes_processados = clientes_dict
            st.success(f"✅ {len(clientes_dict)} clientes encontrados!")
        except Exception as e: st.error(f"❌ Erro: {e}")

if st.session_state.clientes_processados:
    for nome, dados in st.session_state.clientes_processados.items():
        with st.expander(f"📦 CLIENTE: {nome}", expanded=True):
            st.image(dados["img"])
            c1, c2 = st.columns(2)
            with c1:
                if st.button(f"📲 Enviar via WhatsApp: {nome}", key=f"btn_{nome}"):
                    res = enviar_evolution(dados["img"], nome, dados["data_str"], "Próximas 3h")
                    st.success(res) if "✅" in res else st.error(res)
            with c2:
                st.download_button(f"📥 Baixar Excel: {nome}", dados["excel"], f"Escala_{nome.replace('/', '_')}.xlsx", key=f"dl_{nome}")
